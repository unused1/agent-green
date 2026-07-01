"""Build 3 blinded rater copies of the paired-120 human frame (RQ3 revision).

Each rater's copy has their PRIOR scores pre-filled on the 30 reused (correct-case)
rows and blank Likert cells on the 90 new rows. Blinded: model/mode/stratum/reused
withheld; ground-truth label + cwe/cve context shown (safe rows blank cwe/cve, per
the pilot). Reused rows are matched to the prior sheet by exact response text.

Inputs : results/rq3_baseline/fpfn_human_frame_120.csv (master)
         results/rq3_baseline/super49b_zero_rater_sheet v2*.xlsx (prior 3 raters)
Outputs: results/rq3_baseline/fpfn_paired120_rater_<name>.xlsx
"""

import csv, os, sys, glob, re
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
csv.field_size_limit(sys.maxsize)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE = os.path.join(ROOT, "results/rq3_baseline")
MASTER = os.path.join(BASE, "fpfn_human_frame_120.csv")
SCORE_COLS = ["completeness_score", "clarity_score", "actionability_score", "informativeness_score"]
HEADERS = ["sample_id", "source_code", "ground_truth_label", "cwe", "cve_desc",
           "response_text"] + SCORE_COLS + ["rater_notes"]
WIDTHS = {"A": 9, "B": 80, "C": 16, "D": 14, "E": 50, "F": 80,
          "G": 13, "H": 11, "I": 14, "J": 15, "K": 32}


def rater_name(path):
    b = os.path.basename(path).replace(".xlsx", "")
    m = re.split(r"v2", b)[-1]
    return re.sub(r"^[ _-]+", "", m).strip() or "rater"


def load_prior_scores(path):
    """{response_text.strip(): {score_col: value}} from a prior rater sheet."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hdr = [str(h) if h is not None else "" for h in rows[0]]
    col = {h: i for i, h in enumerate(hdr)}
    out = {}
    for r in rows[1:]:
        if not r or all(c is None for c in r):
            continue
        resp = str(r[col["response_text"]] or "").strip()
        out[resp] = {sc: (r[col[sc]] if sc in col and col[sc] < len(r) else None) for sc in SCORE_COLS}
    return out


def main():
    master = list(csv.DictReader(open(MASTER)))
    prior_sheets = sorted(glob.glob(os.path.join(BASE, "super49b_zero_rater_sheet v2*.xlsx")))
    print(f"master rows={len(master)}  prior rater sheets={len(prior_sheets)}")

    for ps in prior_sheets:
        name = rater_name(ps)
        scores = load_prior_scores(ps)
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = f"RQ3 paired-120 {name}"
        hf = PatternFill("solid", fgColor="DDDDDD"); hb = Font(bold=True)
        pf = PatternFill("solid", fgColor="EAF3EA")  # light green = prior (pre-filled)
        for c, h in enumerate(HEADERS, 1):
            cell = ws.cell(1, c, h); cell.font = hb; cell.fill = hf
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        matched = 0
        for r in master:
            safe = r["ground_truth_label"] == "safe"
            reused = r["reused_prior"] == "1"
            sc = scores.get(r["response_text"].strip()) if reused else None
            if reused and sc is not None:
                matched += 1
            vals = [sc.get(k) if sc else "" for k in SCORE_COLS]
            note = "prior (already rated)" if (reused and sc) else ""
            row = [int(r["sample_id"]), r["source_code"], r["ground_truth_label"],
                   "" if safe else r["cwe"], "" if safe else r["cve_desc"],
                   r["response_text"], *vals, note]
            ws.append(row)
            if reused and sc:  # tint the pre-filled rows
                for c in range(1, len(HEADERS) + 1):
                    ws.cell(ws.max_row, c).fill = pf
        for c in range(1, len(HEADERS) + 1):
            for cell in ws[openpyxl.utils.get_column_letter(c)]:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for col, w in WIDTHS.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A2"
        out = os.path.join(BASE, f"fpfn_paired120_rater_{name}.xlsx")
        wb.save(out)
        blank = sum(1 for r in master if r["reused_prior"] != "1")
        print(f"  {name}: pre-filled {matched}/30 reused, {blank} blank to rate -> {os.path.basename(out)}")


if __name__ == "__main__":
    main()
