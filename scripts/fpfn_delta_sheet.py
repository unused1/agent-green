"""Build the FP/FN delta rater sheet: only rows NOT already human-reviewed in v1.

After a frame regeneration (corrected labels), most rows reuse the prior frame's
snippets verbatim (see rq3_generate_fpfn_sample_sets.py --reuse-from). This emits a
sheet containing ONLY the genuinely-new rows so the human re-review is minimal.

Identity is the untruncated (source_code, response_text) pair compared FRAME-vs-FRAME
(not sheet-vs-sheet) to avoid Excel's 32767-char cell truncation causing false
mismatches on long responses. Includes a `parsed_label (TEMP)` column (the parsed
prediction, vulnerable/safe) so the reviewer can check the label against the
response text, matching the v1 sheet format. SAFE rows blank cwe/cve (pilot rule).

Usage:
    python scripts/fpfn_delta_sheet.py
"""

import csv
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

csv.field_size_limit(sys.maxsize)

ROOT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                    "results", "rq3_baseline")
V1_FRAME = os.path.join(ROOT, "fpfn_sample_frame_v1.csv")   # prior reviewed frame
FRAME = os.path.join(ROOT, "fpfn_sample_frame.csv")          # current frame
OUT = os.path.join(ROOT, "fpfn_rater_sheet_v2_delta.xlsx")

HEADERS = ["sample_id", "source_code", "ground_truth_label", "cwe", "cve_desc",
           "response_text", "parsed_label (TEMP)", "completeness_score",
           "clarity_score", "actionability_score", "informativeness_score",
           "rater_notes"]
WIDTHS = {"A": 9, "B": 80, "C": 16, "D": 14, "E": 50, "F": 80, "G": 16,
          "H": 13, "I": 11, "J": 14, "K": 15, "L": 32}


def main():
    v1k = {(r["source_code"].strip(), r["response_text"].strip())
           for r in csv.DictReader(open(V1_FRAME))}
    frame = list(csv.DictReader(open(FRAME)))
    new = [r for r in frame
           if (r["source_code"].strip(), r["response_text"].strip()) not in v1k]
    new.sort(key=lambda r: int(r["sample_id"]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FP-FN delta to review"
    hf = PatternFill("solid", fgColor="DDDDDD"); hb = Font(bold=True)
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(1, c, h); cell.font = hb; cell.fill = hf
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for r in new:
        safe = r["ground_truth_label"] == "safe"
        parsed = "vulnerable" if int(r["prediction"]) == 1 else "safe"
        ws.append([int(r["sample_id"]), r["source_code"], r["ground_truth_label"],
                   "" if safe else r["cwe"], "" if safe else r["cve_desc"],
                   r["response_text"], parsed, "", "", "", "", ""])
    for c in range(1, len(HEADERS) + 1):
        for cell in ws[openpyxl.utils.get_column_letter(c)]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for col, w in WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(OUT)
    print(f"Wrote {OUT} ({len(new)} rows to review; parsed_label at col G)")
    print(f"  sample_ids: {sorted(int(r['sample_id']) for r in new)}")


if __name__ == "__main__":
    main()
