"""Generate the FP/FN explanation-quality sample sets (camera-ready, Item 8).

Design (mirrors the original RQ3 pilot, scaled up):
  - Two model families (independent): Nemotron-Super-49B and Qwen3-30B, SA
    zero-shot. Cross-model is treated as INDEPENDENT replication (RC#7), NOT a
    cross-model intersection.
  - Within each family, sample on the thinking <-> instruct INTERSECTION per
    stratum, so each sampled SNIPPET contributes BOTH a thinking and an instruct
    response on the same code — enabling the paired within-snippet think-vs-
    instruct contrast (the core RQ3 comparison).
        TP: gt=1, think=1, inst=1     FP: gt=0, think=1, inst=1
        TN: gt=0, think=0, inst=0     FN: gt=1, think=0, inst=0
  - N snippets per (family x stratum), default 30 -> 30 x 2 modes = 60 rows.
    Total frame: 2 families x 4 strata x 30 x 2 modes = 480 rows.
  - Labels read from the PATCHED JSONLs (corrected `vuln`), so strata reflect the
    corrected dataset.

Outputs (results/rq3_baseline/):
  fpfn_sample_frame.csv   full frame WITH metadata + source + response. Serves as
                          (a) the LLM-judge input and (b) the unblinding KEY
                          (sample_id -> family/mode/stratum/entry_id/gt).
  fpfn_rater_sheet.xlsx   Human sheet matching the prior pilot format
                          (super49b_zero_rater_sheet): sample_id, source_code,
                          ground_truth_label, cwe, cve_desc, response_text, four
                          blank Likert columns, rater_notes. The ground-truth
                          context (label/cwe/cve) is shown so raters can judge
                          explanation quality against the real vulnerability;
                          model, mode (thinking/instruct), and stratum/correctness
                          are WITHHELD (analysis dimensions) to avoid biasing the
                          ratings. Globally shuffled (seed) so a snippet's two
                          modes are not adjacent. sample_id joins back to the
                          frame CSV for unblinding.

The LLM judge scores the full frame; humans rate a stratified subset for
agreement (per the validated two-stage protocol).
"""

import argparse
import csv
import json
import os
import random
import sys
from collections import defaultdict

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "src"))
from vuln_parser import strip_think_block  # noqa: E402

csv.field_size_limit(sys.maxsize)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CONSOLIDATED = os.path.join(ROOT, "results", "consolidated_performance.csv")
VULN_DATASETS = [
    os.path.join(ROOT, "vuln_database", "VulTrial_486_samples_balanced.jsonl"),
    os.path.join(ROOT, "vuln_database", "VulTrial_384_incremental.jsonl"),
]
OUT_FRAME = os.path.join(ROOT, "results", "rq3_baseline", "fpfn_sample_frame.csv")
OUT_XLSX = os.path.join(ROOT, "results", "rq3_baseline", "fpfn_rater_sheet.xlsx")

SEED = 42

# Family -> (thinking config selector, instruct config selector).
# Each selector is (model_name, mode). Nemotron toggles mode on one checkpoint;
# Qwen ships two checkpoints whose name already encodes the mode.
FAMILIES = {
    "Nemotron-Super-49B": {
        "thinking": ("Nemotron-Super-49B", "thinking"),
        "instruct": ("Nemotron-Super-49B", "instruct"),
    },
    "Qwen3-30B": {
        "thinking": ("Qwen3-30B-A3B-Thinking", "thinking"),
        "instruct": ("Qwen3-30B-A3B-Instruct", "instruct"),
    },
}

# (gt, think_pred, inst_pred) signature per stratum (intersection requirement).
STRATA = {
    "TP": (1, 1, 1),
    "TN": (0, 0, 0),
    "FP": (0, 1, 1),
    "FN": (1, 0, 0),
}


def load_vuln_ds():
    ds = {}
    for p in VULN_DATASETS:
        with open(p) as f:
            for line in f:
                if line.strip():
                    r = json.loads(line)
                    ds[int(r["idx"])] = r
    return ds


def config_index():
    """{(model, mode, 'zero-shot'): [source files]} for SA zero-shot configs."""
    idx = {}
    for r in csv.DictReader(open(CONSOLIDATED)):
        if (r["dataset"] == "VulTrial-870" and r["design"] == "SA"
                and r["prompting"] == "zero-shot"):
            files = [x.strip() for x in r["source_file"].split(";") if x.strip()]
            idx[(r["model"], r["mode"])] = files
    return idx


def load_preds(files):
    """{idx: (gt, pred, reasoning)} deduped by idx; skips/undetermined excluded."""
    out = {}
    for p in files:
        for line in open(p):
            if not line.strip():
                continue
            rec = json.loads(line)
            i = rec.get("idx")
            if i is None or int(i) in out:
                continue
            gt = rec.get("ground_truth", rec.get("target"))
            v = rec.get("vuln")
            try:
                gt = int(gt); v = int(v)
            except (TypeError, ValueError):
                continue
            if v == -1:
                continue
            out[int(i)] = (gt, v, rec.get("reasoning", ""))
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--per-stratum", type=int, default=30)
    args = ap.parse_args()

    vuln_ds = load_vuln_ds()
    cfg = config_index()
    rng = random.Random(SEED)

    rows = []
    print(f"Sampling {args.per_stratum} snippets per (family x stratum), seed={SEED}\n")
    for family, modes in FAMILIES.items():
        tk_model, tk_mode = modes["thinking"]
        in_model, in_mode = modes["instruct"]
        think = load_preds(cfg[(tk_model, tk_mode)])
        inst = load_preds(cfg[(in_model, in_mode)])
        common = set(think) & set(inst)

        for stratum, (gt_req, tk_req, in_req) in STRATA.items():
            pool = [i for i in common
                    if think[i][0] == gt_req and think[i][1] == tk_req
                    and inst[i][0] == gt_req and inst[i][1] == in_req]
            pool.sort()
            rng.shuffle(pool)
            take = pool[:args.per_stratum]
            if len(pool) < args.per_stratum:
                print(f"  WARN {family} {stratum}: only {len(pool)} in intersection")
            for i in take:
                vrec = vuln_ds.get(i, {})
                base = {
                    "family": family, "stratum": stratum, "entry_id": i,
                    "snippet_pair_id": f"{family}:{stratum}:{i}",
                    "ground_truth": gt_req,
                    "ground_truth_label": "vulnerable" if gt_req == 1 else "safe",
                    "cwe": str(vrec.get("cwe", "")),
                    "cve_desc": str(vrec.get("cve_desc", "")),
                    "source_code": vrec.get("func", ""),
                }
                rows.append({**base, "model": tk_model, "mode": "thinking",
                             "prediction": think[i][1],
                             "response_text": strip_think_block(think[i][2])})
                rows.append({**base, "model": in_model, "mode": "instruct",
                             "prediction": inst[i][1],
                             "response_text": strip_think_block(inst[i][2])})
            print(f"  {family:20s} {stratum}: {len(take)} snippets x2 modes = {len(take)*2} rows")

    # Global shuffle + assign sample_id (so a snippet's two modes are not adjacent
    # and order leaks nothing).
    rng2 = random.Random(SEED)
    rng2.shuffle(rows)
    for sid, r in enumerate(rows, 1):
        r["sample_id"] = sid

    # Frame CSV = LLM-judge input + unblinding key
    fields = ["sample_id", "family", "model", "mode", "stratum", "entry_id",
              "snippet_pair_id", "ground_truth", "prediction", "ground_truth_label",
              "cwe", "cve_desc", "source_code", "response_text"]
    with open(OUT_FRAME, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)
    print(f"\nWrote {OUT_FRAME} ({len(rows)} rows)")

    # Blind rater XLSX: sample_id + code + response + blank Likert only
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print("openpyxl not available; CSV frame written, skipping XLSX.")
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FP-FN rater sheet"
    headers = ["sample_id", "source_code", "ground_truth_label", "cwe", "cve_desc",
               "response_text", "completeness_score", "clarity_score",
               "actionability_score", "informativeness_score", "rater_notes"]
    hf = PatternFill("solid", fgColor="DDDDDD"); hb = Font(bold=True)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h); cell.font = hb; cell.fill = hf
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    for r in rows:  # already shuffled; sample_id is the key into the frame CSV
        # Blank cwe/cve for SAFE rows (match the prior pilot): the benign member of
        # a PrimeVul pair inherits its mate's CVE, which describes a vulnerability
        # the patched code no longer contains — showing it would mislead the rater.
        safe = r["ground_truth_label"] == "safe"
        cwe = "" if safe else r["cwe"]
        cve = "" if safe else r["cve_desc"]
        ws.append([r["sample_id"], r["source_code"], r["ground_truth_label"],
                   cwe, cve, r["response_text"], "", "", "", "", ""])
    for c in range(1, len(headers) + 1):
        for cell in ws[openpyxl.utils.get_column_letter(c)]:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = {"A": 9, "B": 80, "C": 16, "D": 14, "E": 50, "F": 80,
              "G": 13, "H": 11, "I": 14, "J": 15, "K": 32}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)
    print(f"Wrote {OUT_XLSX} ({len(rows)} rows; mode/stratum withheld; "
          f"key=sample_id in frame CSV)")

    from collections import Counter
    dist = Counter((r["family"], r["stratum"], r["mode"]) for r in rows)
    print("\nPer (family, stratum, mode):")
    for k in sorted(dist):
        print(f"  {k}: {dist[k]}")


if __name__ == "__main__":
    main()
