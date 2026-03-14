"""
Compute inter-rater reliability (IRR) metrics for RQ3 Phase A human ratings.

Reads rater xlsx files (Shane, HS, and optionally more), joins to master
metadata via source_code + response_text hashing, computes per-dimension
ICC, Spearman rho, Cohen's weighted kappa, flags disagreements > 1 point,
and produces consensus scores.

Outputs:
    results/rq3_baseline/irr_summary.csv
    results/rq3_baseline/irr_disagreements.csv
    results/rq3_baseline/super49b_zero_consensus_scores.csv
"""

import argparse
import csv
import hashlib
import os
import sys

import numpy as np
import openpyxl
import pandas as pd
import pingouin as pg
from scipy.stats import spearmanr
from sklearn.metrics import cohen_kappa_score

csv.field_size_limit(sys.maxsize)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "results", "rq3_baseline")
MASTER_CSV = os.path.join(OUTPUT_DIR, "super49b_zero_human_rating_set.csv")
RATER_SHEET_CSV = os.path.join(OUTPUT_DIR, "super49b_zero_rater_sheet.csv")

DIMENSIONS = ["completeness", "clarity", "actionability", "informativeness"]
SCORE_COLS = [f"{d}_score" for d in DIMENSIONS]

# Default rater files (can be extended via CLI)
DEFAULT_RATERS = {
    "Shane": os.path.join(OUTPUT_DIR, "super49b_zero_rater_sheet v2 - Shane.xlsx"),
    "HS": os.path.join(OUTPUT_DIR, "super49b_zero_rater_sheet v2-HS.xlsx"),
}


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
def load_rater_xlsx(path: str) -> pd.DataFrame:
    """Load a rater xlsx file. Returns DataFrame with sample_id and 4 score columns."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    wb.close()
    df = pd.DataFrame(rows)
    # Ensure score columns are numeric
    for col in SCORE_COLS:
        df[col] = pd.to_numeric(df[col], errors="coerce")
    df["sample_id"] = pd.to_numeric(df["sample_id"], errors="coerce").astype(int)
    return df


def build_master_mapping() -> dict:
    """Build mapping from hash(source_code + rater_response_text) -> metadata dict.

    Uses the master human_rating_set.csv which contains entry_id, response_id,
    ground_truth, etc.
    """
    mapping = {}
    with open(MASTER_CSV, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = hashlib.md5(
                (row["source_code"] + row["rater_response_text"]).encode()
            ).hexdigest()
            mapping[key] = {
                "entry_id": int(row["entry_id"]),
                "response_id": row["response_id"],
                "ground_truth": int(row["ground_truth"]),
                "ground_truth_label": row["ground_truth_label"],
                "snippet_id": int(row["sample_id"]),
            }
    return mapping


def build_rater_to_master_map() -> dict:
    """Map rater_sheet sample_id (1-30) -> master metadata."""
    master_map = build_master_mapping()

    rater_map = {}
    with open(RATER_SHEET_CSV, newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = hashlib.md5(
                (row["source_code"] + row["response_text"]).encode()
            ).hexdigest()
            sid = int(row["sample_id"])
            if key in master_map:
                rater_map[sid] = master_map[key]
            else:
                print(f"WARNING: rater_sheet sample_id {sid} has no match in master")
    return rater_map


# ---------------------------------------------------------------------------
# IRR computation
# ---------------------------------------------------------------------------
def compute_icc(scores_df: pd.DataFrame, dimension: str, rater_names: list) -> dict:
    """Compute ICC(2,1) for a single dimension using pingouin.

    Args:
        scores_df: DataFrame with columns [sample_id, rater, score]
        dimension: Name of the dimension (for labeling)
        rater_names: List of rater names

    Returns:
        dict with ICC type, value, CI, p-value
    """
    n_raters = len(rater_names)
    # Use ICC(2,1) for 2 raters, ICC(2,k) for 3+
    icc_type = "ICC2" if n_raters == 2 else "ICC2k"

    try:
        icc_result = pg.intraclass_corr(
            data=scores_df,
            targets="sample_id",
            raters="rater",
            ratings="score",
        )
        # ICC2 is row index 1 (ICC2,1), ICC2k is row index 4 (ICC2,k)
        row_idx = 1 if icc_type == "ICC2" else 4
        icc_row = icc_result.iloc[row_idx]
        # pingouin uses 'CI95' (list of [lower, upper])
        ci = icc_row.get("CI95", icc_row.get("CI95%", [np.nan, np.nan]))
        return {
            "icc_type": icc_row["Type"],
            "icc_value": round(icc_row["ICC"], 4),
            "icc_ci_lower": round(ci[0], 4),
            "icc_ci_upper": round(ci[1], 4),
            "icc_pvalue": round(icc_row["pval"], 6),
        }
    except Exception as e:
        print(f"  WARNING: ICC computation failed for {dimension}: {e}")
        return {
            "icc_type": icc_type,
            "icc_value": np.nan,
            "icc_ci_lower": np.nan,
            "icc_ci_upper": np.nan,
            "icc_pvalue": np.nan,
        }


def compute_irr_metrics(
    rater_scores: dict, rater_names: list
) -> pd.DataFrame:
    """Compute per-dimension IRR metrics across all raters.

    Args:
        rater_scores: {rater_name: DataFrame with sample_id and score columns}
        rater_names: ordered list of rater names

    Returns:
        DataFrame with one row per dimension and IRR statistics
    """
    results = []

    for dim in DIMENSIONS:
        col = f"{dim}_score"

        # Build long-format DataFrame for ICC
        long_rows = []
        for rname in rater_names:
            df = rater_scores[rname]
            for _, row in df.iterrows():
                long_rows.append({
                    "sample_id": row["sample_id"],
                    "rater": rname,
                    "score": row[col],
                })
        long_df = pd.DataFrame(long_rows)

        # ICC
        icc_info = compute_icc(long_df, dim, rater_names)

        # For pairwise metrics (2 raters), compute Spearman and weighted kappa
        if len(rater_names) == 2:
            r1 = rater_scores[rater_names[0]].sort_values("sample_id")[col].values
            r2 = rater_scores[rater_names[1]].sort_values("sample_id")[col].values

            rho, rho_p = spearmanr(r1, r2)
            wkappa = cohen_kappa_score(r1, r2, weights="quadratic")

            # Agreement statistics
            diffs = r1 - r2
            abs_diffs = np.abs(diffs)
            mean_abs_diff = np.mean(abs_diffs)
            pct_perfect = np.mean(abs_diffs == 0) * 100
            pct_within1 = np.mean(abs_diffs <= 1) * 100
            mean_signed_diff = np.mean(diffs)  # positive = rater1 higher
        else:
            # For 3+ raters, compute average pairwise metrics
            rhos, kappas = [], []
            for i in range(len(rater_names)):
                for j in range(i + 1, len(rater_names)):
                    r_i = rater_scores[rater_names[i]].sort_values("sample_id")[col].values
                    r_j = rater_scores[rater_names[j]].sort_values("sample_id")[col].values
                    rho_ij, _ = spearmanr(r_i, r_j)
                    rhos.append(rho_ij)
                    kappas.append(cohen_kappa_score(r_i, r_j, weights="quadratic"))
            rho = np.mean(rhos)
            rho_p = np.nan  # no single p-value for averaged rho
            wkappa = np.mean(kappas)
            # Agreement stats using all pairwise diffs
            all_diffs = []
            for i in range(len(rater_names)):
                for j in range(i + 1, len(rater_names)):
                    r_i = rater_scores[rater_names[i]].sort_values("sample_id")[col].values
                    r_j = rater_scores[rater_names[j]].sort_values("sample_id")[col].values
                    all_diffs.extend(r_i - r_j)
            all_diffs = np.array(all_diffs)
            abs_diffs = np.abs(all_diffs)
            mean_abs_diff = np.mean(abs_diffs)
            pct_perfect = np.mean(abs_diffs == 0) * 100
            pct_within1 = np.mean(abs_diffs <= 1) * 100
            mean_signed_diff = np.mean(all_diffs)

        # Rater means
        rater_means = {}
        for rname in rater_names:
            rater_means[f"mean_{rname}"] = round(
                rater_scores[rname][col].mean(), 2
            )

        row = {
            "dimension": dim,
            **icc_info,
            "spearman_rho": round(rho, 4),
            "spearman_pvalue": round(rho_p, 6) if not np.isnan(rho_p) else np.nan,
            "weighted_kappa": round(wkappa, 4),
            "mean_abs_diff": round(mean_abs_diff, 3),
            "pct_perfect_agreement": round(pct_perfect, 1),
            "pct_within_1": round(pct_within1, 1),
            "mean_signed_diff": round(mean_signed_diff, 3),
            **rater_means,
        }
        results.append(row)

    return pd.DataFrame(results)


def find_disagreements(
    rater_scores: dict, rater_names: list, threshold: int = 1
) -> pd.DataFrame:
    """Flag samples where any dimension has |diff| > threshold between any rater pair."""
    disagreements = []

    # Build merged DataFrame
    base = rater_scores[rater_names[0]].sort_values("sample_id").reset_index(drop=True)
    for rname in rater_names[1:]:
        other = rater_scores[rname].sort_values("sample_id").reset_index(drop=True)
        for dim in DIMENSIONS:
            col = f"{dim}_score"
            base[f"{col}_{rname}"] = other[col].values

    for _, row in base.iterrows():
        sid = int(row["sample_id"])
        flagged_dims = []
        for dim in DIMENSIONS:
            col = f"{dim}_score"
            scores = [row[col]]  # first rater
            for rname in rater_names[1:]:
                scores.append(row[f"{col}_{rname}"])
            max_diff = max(scores) - min(scores)
            if max_diff > threshold:
                flagged_dims.append(dim)

        if flagged_dims:
            d = {"sample_id": sid}
            for dim in DIMENSIONS:
                col = f"{dim}_score"
                for i, rname in enumerate(rater_names):
                    if i == 0:
                        d[f"{dim}_{rname}"] = int(row[col])
                    else:
                        d[f"{dim}_{rname}"] = int(row[f"{col}_{rname}"])
                # Compute diff for this dimension
                vals = [d[f"{dim}_{rname}"] for rname in rater_names]
                d[f"{dim}_diff"] = max(vals) - min(vals)
            d["flagged_dimensions"] = "; ".join(flagged_dims)
            disagreements.append(d)

    return pd.DataFrame(disagreements)


def compute_consensus(
    rater_scores: dict, rater_names: list, rater_to_master: dict
) -> pd.DataFrame:
    """Compute consensus scores (average) and attach master metadata."""
    rows = []

    # Get sorted sample_ids
    sample_ids = sorted(rater_scores[rater_names[0]]["sample_id"].unique())

    for sid in sample_ids:
        row = {"sample_id": sid}

        # Master metadata
        meta = rater_to_master.get(sid, {})
        row["entry_id"] = meta.get("entry_id", "")
        row["response_id"] = meta.get("response_id", "")
        row["ground_truth"] = meta.get("ground_truth", "")
        row["ground_truth_label"] = meta.get("ground_truth_label", "")
        row["snippet_id"] = meta.get("snippet_id", "")

        # Derive stratum
        resp_id = meta.get("response_id", "")
        gt = meta.get("ground_truth", "")
        if resp_id and gt != "":
            mode = "think" if resp_id == "think" else "inst"
            label = "TP" if int(gt) == 1 else "TN"
            row["stratum"] = f"{mode}-{label}"
        else:
            row["stratum"] = ""

        # Per-rater scores and consensus
        for dim in DIMENSIONS:
            col = f"{dim}_score"
            rater_vals = []
            for rname in rater_names:
                df = rater_scores[rname]
                val = df.loc[df["sample_id"] == sid, col].values[0]
                row[f"{dim}_{rname}"] = int(val)
                rater_vals.append(val)
            row[f"{dim}_consensus"] = round(np.mean(rater_vals), 2)
            row[f"{dim}_diff"] = int(max(rater_vals) - min(rater_vals))

        rows.append(row)

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Interpretive helpers
# ---------------------------------------------------------------------------
def icc_interpretation(icc_val: float) -> str:
    """Interpret ICC value per Koo & Li (2016) guidelines."""
    if np.isnan(icc_val):
        return "N/A"
    if icc_val < 0.50:
        return "poor"
    elif icc_val < 0.75:
        return "moderate"
    elif icc_val < 0.90:
        return "good"
    else:
        return "excellent"


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Compute inter-rater reliability for RQ3 Phase A ratings"
    )
    parser.add_argument(
        "--raters", nargs="*",
        help="Additional rater xlsx files as name:path pairs (e.g., R3:/path/to/file.xlsx)"
    )
    args = parser.parse_args()

    # Build rater file dict
    rater_files = dict(DEFAULT_RATERS)
    if args.raters:
        for spec in args.raters:
            name, path = spec.split(":", 1)
            rater_files[name] = path

    rater_names = list(rater_files.keys())
    print(f"Raters: {rater_names}")
    print(f"Dimensions: {DIMENSIONS}")
    print()

    # Load rater scores
    rater_scores = {}
    for rname, path in rater_files.items():
        if not os.path.exists(path):
            sys.exit(f"ERROR: Rater file not found: {path}")
        df = load_rater_xlsx(path)
        print(f"Loaded {rname}: {len(df)} rows, "
              f"score range {df[SCORE_COLS].min().min()}-{df[SCORE_COLS].max().max()}")
        # Validate no missing scores
        missing = df[SCORE_COLS].isna().sum().sum()
        if missing > 0:
            print(f"  WARNING: {missing} missing scores in {rname}")
        rater_scores[rname] = df

    print()

    # Build master mapping
    print("Building rater_sheet -> master metadata mapping...")
    rater_to_master = build_rater_to_master_map()
    print(f"  Mapped {len(rater_to_master)} / 30 sample_ids")
    print()

    # --- IRR metrics ---
    print("=" * 70)
    print("INTER-RATER RELIABILITY METRICS")
    print("=" * 70)
    irr_df = compute_irr_metrics(rater_scores, rater_names)

    for _, row in irr_df.iterrows():
        dim = row["dimension"]
        interp = icc_interpretation(row["icc_value"])
        print(f"\n  {dim.upper()}")
        print(f"    ICC({row['icc_type']}):    {row['icc_value']:.4f} "
              f"[{row['icc_ci_lower']:.4f}, {row['icc_ci_upper']:.4f}] "
              f"p={row['icc_pvalue']:.4f} — {interp}")
        print(f"    Spearman ρ:    {row['spearman_rho']:.4f} "
              f"(p={row['spearman_pvalue']:.4f})")
        print(f"    Weighted κ:    {row['weighted_kappa']:.4f}")
        print(f"    Mean |diff|:   {row['mean_abs_diff']:.2f}")
        print(f"    Perfect agmt:  {row['pct_perfect_agreement']:.1f}%")
        print(f"    Within 1 pt:   {row['pct_within_1']:.1f}%")
        bias_dir = f"{rater_names[0]} higher" if row["mean_signed_diff"] > 0 else f"{rater_names[1]} higher"
        print(f"    Signed diff:   {row['mean_signed_diff']:+.2f} ({bias_dir})")
        for rname in rater_names:
            print(f"    Mean ({rname:>6s}): {row[f'mean_{rname}']:.2f}")

    # Save IRR summary
    irr_path = os.path.join(OUTPUT_DIR, "irr_summary.csv")
    irr_df.to_csv(irr_path, index=False)
    print(f"\n  Saved: {irr_path}")

    # --- Disagreements ---
    print()
    print("=" * 70)
    print("DISAGREEMENTS (|diff| > 1 on any dimension)")
    print("=" * 70)
    disagree_df = find_disagreements(rater_scores, rater_names, threshold=1)

    if len(disagree_df) == 0:
        print("  No disagreements > 1 point found.")
    else:
        print(f"  {len(disagree_df)} samples with disagreements > 1 point:")
        for _, row in disagree_df.iterrows():
            sid = int(row["sample_id"])
            meta = rater_to_master.get(sid, {})
            eid = meta.get("entry_id", "?")
            rid = meta.get("response_id", "?")
            flagged = row["flagged_dimensions"]
            print(f"\n    sample_id={sid} (entry={eid}, resp={rid})")
            for dim in DIMENSIONS:
                scores = [str(int(row[f"{dim}_{rn}"])) for rn in rater_names]
                diff = int(row[f"{dim}_diff"])
                flag = " ***" if diff > 1 else ""
                print(f"      {dim:18s}: {' / '.join(scores)}  (diff={diff}){flag}")

    disagree_path = os.path.join(OUTPUT_DIR, "irr_disagreements.csv")
    disagree_df.to_csv(disagree_path, index=False)
    print(f"\n  Saved: {disagree_path}")

    # --- Consensus scores ---
    print()
    print("=" * 70)
    print("CONSENSUS SCORES")
    print("=" * 70)
    consensus_df = compute_consensus(rater_scores, rater_names, rater_to_master)

    # Check for low-ICC dimensions
    low_icc_dims = []
    for _, row in irr_df.iterrows():
        if row["icc_value"] < 0.5:
            low_icc_dims.append(row["dimension"])
    if low_icc_dims:
        print(f"  WARNING: Low ICC (< 0.5) on: {', '.join(low_icc_dims)}")
        print("  → Consider discussion-based resolution for these dimensions")

    # Summary statistics
    print(f"\n  Consensus score summary (N={len(consensus_df)}):")
    for dim in DIMENSIONS:
        col = f"{dim}_consensus"
        vals = consensus_df[col]
        print(f"    {dim:18s}: mean={vals.mean():.2f}, "
              f"sd={vals.std():.2f}, range=[{vals.min():.1f}, {vals.max():.1f}]")

    # Per-stratum summary
    print(f"\n  Per-stratum means:")
    for stratum in sorted(consensus_df["stratum"].unique()):
        sub = consensus_df[consensus_df["stratum"] == stratum]
        means = [f"{dim[:4]}={sub[f'{dim}_consensus'].mean():.2f}" for dim in DIMENSIONS]
        print(f"    {stratum:10s} (n={len(sub):2d}): {', '.join(means)}")

    consensus_path = os.path.join(OUTPUT_DIR, "super49b_zero_consensus_scores.csv")
    consensus_df.to_csv(consensus_path, index=False)
    print(f"\n  Saved: {consensus_path}")

    # --- Overall summary ---
    print()
    print("=" * 70)
    print("INTERPRETATION SUMMARY")
    print("=" * 70)
    all_good = True
    for _, row in irr_df.iterrows():
        interp = icc_interpretation(row["icc_value"])
        icc_val = row["icc_value"]
        status = "✓" if (not np.isnan(icc_val) and icc_val >= 0.5) else "✗"
        if np.isnan(icc_val) or icc_val < 0.5:
            all_good = False
        print(f"  {status} {row['dimension']:18s}: ICC={row['icc_value']:.3f} ({interp})")

    if all_good:
        print("\n  All dimensions have ICC >= 0.5 — averaging is appropriate for consensus.")
    else:
        print("\n  Some dimensions have ICC < 0.5 — discussion-based resolution recommended.")

    print(f"\n  Total consensus rows: {len(consensus_df)}")
    print(f"  Disagreements > 1pt: {len(disagree_df)}")


if __name__ == "__main__":
    main()
