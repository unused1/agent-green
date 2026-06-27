"""
Export the qualitative FP/FN cases to an XLSX in the same shape as
`super49b_zero_rater_sheet.xlsx`, with extra columns for the LLM judge
scores + justifications and a flag marking the cases recommended for the
ASE 2026 rebuttal.

Output: results/rq3_baseline/rq3_b5_qualitative_cases_rater_sheet.xlsx
"""

import csv
import re
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

csv.field_size_limit(sys.maxsize)


def _strip_think_tags(text: str) -> str:
    """Remove <think>...</think> blocks from response (matches the
    same processing applied in the Phase A rater sheet and to the LLM
    judge inputs)."""
    text = re.sub(r"<think>.*?</think>\s*", "", text or "", flags=re.DOTALL)
    if "</think>" in text:
        text = text.split("</think>", 1)[1].strip()
    return text


def _prepare_rater_text(text: str, response_id: str) -> str:
    """Strip think tags for thinking-mode responses; leave instruct as-is."""
    return _strip_think_tags(text) if response_id == "think" else (text or "")

PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = PROJECT_ROOT / "results" / "rq3_baseline"
OUT_PATH = OUT_DIR / "rq3_b5_qualitative_cases_rater_sheet.xlsx"

DIMENSIONS = ["completeness", "clarity", "actionability", "informativeness"]

MODEL_REGISTRY = {
    "super49b": {
        "display_name": "Nemotron-Super-49B",
        "rating_set": "super49b_zero_incorrect_rating_set.csv",
        "judged":     "super49b_zero_incorrect_llm_judged_opus-4-6_zeroshot.csv",
    },
    "qwen30b": {
        "display_name": "Qwen3-30B-A3B",
        "rating_set": "qwen30b_zero_incorrect_rating_set.csv",
        "judged":     "qwen30b_zero_incorrect_llm_judged_opus-4-6_zeroshot.csv",
    },
}

# Cases recommended for the rebuttal (model_key, entry_id, response_id).
# Updated after the 2026-06-08 parser-mismatch filter regenerated the pools.
# 4 cases covering both models × both error types, each illustrating the
# "confident but causally-disconnected" pattern with a real CVE description.
REBUTTAL_RECOMMENDED = {
    # Super-49B FN — instruct — CWE-522, BIND MITM cache-poisoning
    ("super49b", 195388, "inst"),
    # Super-49B FP — thinking — CWE-369, TensorFlow divide-by-zero
    ("super49b", 251946, "think"),
    # Qwen3-30B FN — instruct — CWE-362, OpenSSL race condition (high-profile CVE)
    ("qwen30b",  216515, "inst"),
    # Qwen3-30B FP — instruct — CWE-703, TensorFlow exception-handling false alarm
    ("qwen30b",  259732, "inst"),
}


def load_rating_set(path: Path) -> dict:
    out = {}
    with open(path, newline="") as f:
        for row in csv.DictReader(f):
            out[(int(row["entry_id"]), row["response_id"])] = row
    return out


def load_judged(path: Path) -> dict:
    out = {}
    with open(path, newline="") as f:
        for row in csv.DictReader(f):
            out[(int(row["entry_id"]), row["response_id"])] = row
    return out


def plausibility_disconnect(judged_row: dict) -> int:
    s = lambda d: int(judged_row.get(f"{d}_score", 0) or 0)
    return (s("actionability") + s("clarity")) - (s("completeness") + s("informativeness"))


def collect_cases() -> list:
    """Return all 8 stratified cases ordered by (model, error, mode)."""
    cases = []
    for model_key, cfg in MODEL_REGISTRY.items():
        rating_set = load_rating_set(OUT_DIR / cfg["rating_set"])
        judged = load_judged(OUT_DIR / cfg["judged"])
        # Group by (error_type, response_id) so we pick top-disconnect from each
        from collections import defaultdict
        by_strat = defaultdict(list)
        for key, rs_row in rating_set.items():
            if key not in judged:
                continue
            j_row = judged[key]
            err = "FP" if rs_row["stratum"].endswith("FP") else "FN"
            by_strat[(err, key[1])].append({
                "model_key": model_key,
                "model_name": cfg["display_name"],
                "entry_id": key[0],
                "response_id": key[1],
                "error_type": err,
                "rating_set": rs_row,
                "judged": j_row,
                "disconnect": plausibility_disconnect(j_row),
            })
        for k in sorted(by_strat.keys()):
            by_strat[k].sort(key=lambda c: -c["disconnect"])
            cases.append(by_strat[k][0])
    return cases


def write_xlsx(cases: list):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Qualitative cases"

    headers = [
        "sample_id",
        "recommended_for_rebuttal",
        "model_name",
        "error_type",
        "response_mode",
        "entry_id",
        "ground_truth_label",
        "cwe",
        "cve_desc",
        "source_code",
        "response_text",
        "completeness_score",
        "clarity_score",
        "actionability_score",
        "informativeness_score",
        "plausibility_disconnect",
        "completeness_judge_justification",
        "clarity_judge_justification",
        "actionability_judge_justification",
        "informativeness_judge_justification",
        "rater_notes",
    ]

    # Header row
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_font = Font(bold=True)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Data rows
    rec_fill = PatternFill("solid", fgColor="FFF2CC")  # light yellow for recommended
    for sample_id, case in enumerate(cases, 1):
        rs = case["rating_set"]
        j = case["judged"]
        is_rec = (case["model_key"], case["entry_id"], case["response_id"]) in REBUTTAL_RECOMMENDED
        row_values = [
            sample_id,
            "Yes" if is_rec else "No",
            case["model_name"],
            case["error_type"],
            "thinking" if case["response_id"] == "think" else "instruct",
            case["entry_id"],
            rs.get("ground_truth_label", ""),
            rs.get("cwe", ""),
            rs.get("cve_desc", ""),
            rs.get("source_code", ""),
            _prepare_rater_text(rs.get("response_text", ""), case["response_id"]),
            int(j.get("completeness_score") or 0),
            int(j.get("clarity_score") or 0),
            int(j.get("actionability_score") or 0),
            int(j.get("informativeness_score") or 0),
            case["disconnect"],
            j.get("completeness_justification", ""),
            j.get("clarity_justification", ""),
            j.get("actionability_justification", ""),
            j.get("informativeness_justification", ""),
            "",  # empty rater_notes column
        ]
        for c, val in enumerate(row_values, 1):
            cell = ws.cell(sample_id + 1, c, val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if is_rec:
                cell.fill = rec_fill

    # Column widths (tuned for readability)
    widths = {
        "A": 10, "B": 18, "C": 22, "D": 6, "E": 12, "F": 12,
        "G": 14, "H": 14, "I": 40, "J": 70, "K": 70,
        "L": 14, "M": 12, "N": 14, "O": 16, "P": 14,
        "Q": 50, "R": 50, "S": 50, "T": 50, "U": 40,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Freeze header row + first two cols
    ws.freeze_panes = "C2"

    # Hide source_code and response_text by default? No — keep visible for raters.

    # Brief instructions sheet
    info = wb.create_sheet("README")
    info["A1"] = "Qualitative FP/FN case digest for the ASE 2026 rebuttal (Reviewer C #7 + B5)"
    info["A1"].font = Font(bold=True, size=13)
    info["A3"] = (
        "Each row is one (model × error_type × response_mode) case from the "
        "SA zero-shot incorrect-intersection pool, selected by highest "
        "'plausibility-disconnect' score: actionability + clarity − completeness − informativeness."
    )
    info["A4"] = (
        "Yellow-highlighted rows are the cases recommended for the rebuttal narrative "
        "(2 standalone + 1 paired Super-49B FN). You can override the selection by "
        "editing the 'recommended_for_rebuttal' column."
    )
    info["A6"] = "Score interpretation: 1 = poor, 5 = excellent (LLM judge: Claude Opus 4.6 zero-shot)."
    info["A7"] = "Plausibility-disconnect interpretation: higher = sounds clear/actionable but thin on substance."
    for r in range(1, 8):
        info.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")
    info.column_dimensions["A"].width = 110

    wb.save(OUT_PATH)
    print(f"Wrote {OUT_PATH}")
    print(f"  {len(cases)} cases total; {len(REBUTTAL_RECOMMENDED)} flagged as recommended for rebuttal")


def main():
    cases = collect_cases()
    write_xlsx(cases)


if __name__ == "__main__":
    main()
